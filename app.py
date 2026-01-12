import streamlit as st
import anthropic
import base64
import os
import io
import re
from pathlib import Path
from dotenv import load_dotenv
import fitz  # PyMuPDF para leer PDFs
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Cargar variables de entorno (local) o secrets (Streamlit Cloud)
load_dotenv()

def get_secret(key):
    """Obtiene un secret desde Streamlit Cloud o .env local"""
    try:
        return st.secrets[key]
    except:
        return os.getenv(key)

def check_login():
    """Verifica las credenciales de login"""
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.title("🔐 Iniciar Sesión")
    st.markdown("---")

    with st.form("login_form"):
        email = st.text_input("Email", placeholder="usuario@ejemplo.com")
        password = st.text_input("Contraseña", type="password")
        submit = st.form_submit_button("Entrar", use_container_width=True)

        if submit:
            valid_user = get_secret("LOGIN_USER")
            valid_password = get_secret("LOGIN_PASSWORD")

            if email == valid_user and password == valid_password:
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("❌ Credenciales incorrectas")

    return False

def get_convenios_disponibles():
    """Lista los convenios PDF disponibles en la carpeta"""
    carpeta = Path(__file__).parent
    convenios = list(carpeta.glob("*.pdf"))
    return [c for c in convenios if "uploaded_" not in c.name]

def pdf_to_images(pdf_bytes):
    """Convierte un PDF a lista de imágenes en base64"""
    images = []
    pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")

    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        mat = fitz.Matrix(2, 2)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        img_base64 = base64.standard_b64encode(img_bytes).decode("utf-8")
        images.append(img_base64)

    pdf_document.close()
    return images

def extract_text_from_pdf(pdf_bytes):
    """Extrae texto de un PDF"""
    text = ""
    pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")

    for page_num in range(len(pdf_document)):
        page = pdf_document.load_page(page_num)
        text += page.get_text() + "\n\n"

    pdf_document.close()
    return text

def extract_convenio_from_image(client, image_bytes, image_type, detailed=True):
    """Extrae información del convenio desde una imagen usando Claude"""
    img_base64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    media_type = f"image/{image_type}" if image_type != "jpg" else "image/jpeg"

    if detailed:
        prompt_text = """Extrae TODA la información de este convenio colectivo o tabla salarial.

Incluye:
- Tablas salariales con todas las categorías y salarios
- Complementos (transporte, nocturnidad, festividad, etc.)
- Antigüedad (trienios, porcentajes)
- Pagas extras
- Jornada laboral
- Cualquier otro dato relevante para calcular costes de personal

Transcribe los datos de forma estructurada y completa."""
    else:
        prompt_text = """Analiza brevemente esta página. ¿Contiene alguno de estos elementos?
- Tablas salariales o retribuciones
- Complementos salariales (transporte, nocturnidad, etc.)
- Información sobre antigüedad/trienios
- Pagas extraordinarias
- Jornada laboral

Responde SOLO con: "RELEVANTE: [motivo breve]" o "NO RELEVANTE"."""

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096 if detailed else 200,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": img_base64,
                        }
                    },
                    {
                        "type": "text",
                        "text": prompt_text
                    }
                ]
            }
        ]
    )

    return response.content[0].text

def identify_relevant_pages(client, images, progress_callback=None):
    """Identifica qué páginas del convenio contienen información relevante"""
    relevant_pages = []
    total_pages = len(images)

    for i, img_base64 in enumerate(images):
        if progress_callback:
            progress_callback(i + 1, total_pages)

        img_bytes = base64.standard_b64decode(img_base64)
        result = extract_convenio_from_image(client, img_bytes, "png", detailed=False)

        if "RELEVANTE" in result.upper():
            relevant_pages.append(i)

    return relevant_pages

def extract_convenio_from_file(client, file_bytes, file_type, is_image, progress_placeholder=None):
    """Extrae información del convenio desde PDF o imagen con extracción inteligente"""
    if is_image:
        return extract_convenio_from_image(client, file_bytes, file_type, detailed=True)

    # Para PDFs, intentar extraer texto primero
    text = extract_text_from_pdf(file_bytes)

    # Si hay suficiente texto, usarlo directamente
    if len(text.strip()) >= 500:
        return text

    # PDF escaneado - usar extracción inteligente
    images = pdf_to_images(file_bytes)
    total_pages = len(images)

    if progress_placeholder:
        progress_placeholder.info(f"📄 Convenio de {total_pages} páginas detectado. Analizando estructura...")

    # Si son pocas páginas, procesar todas
    if total_pages <= 20:
        all_text = ""
        for i, img_base64 in enumerate(images):
            if progress_placeholder:
                progress_placeholder.info(f"📄 Procesando página {i+1}/{total_pages}...")
            img_bytes = base64.standard_b64decode(img_base64)
            page_text = extract_convenio_from_image(client, img_bytes, "png", detailed=True)
            all_text += f"\n--- PÁGINA {i+1} ---\n{page_text}\n"
        return all_text

    # Para convenios largos: extracción inteligente en 2 fases
    if progress_placeholder:
        progress_placeholder.info(f"🔍 Fase 1: Escaneando {total_pages} páginas para identificar tablas salariales...")

    def update_progress(current, total):
        if progress_placeholder:
            progress_placeholder.info(f"🔍 Fase 1: Escaneando página {current}/{total}...")

    # Fase 1: Identificar páginas relevantes
    relevant_pages = identify_relevant_pages(client, images, update_progress)

    if not relevant_pages:
        # Si no encontró páginas relevantes, usar las primeras 30 páginas
        if progress_placeholder:
            progress_placeholder.warning("⚠️ No se identificaron páginas con tablas. Procesando primeras 30 páginas...")
        relevant_pages = list(range(min(30, total_pages)))
    else:
        if progress_placeholder:
            progress_placeholder.success(f"✅ Encontradas {len(relevant_pages)} páginas con información salarial")

    # Fase 2: Extraer contenido detallado de páginas relevantes
    if progress_placeholder:
        progress_placeholder.info(f"📊 Fase 2: Extrayendo datos de {len(relevant_pages)} páginas relevantes...")

    all_text = ""
    for idx, page_num in enumerate(relevant_pages):
        if progress_placeholder:
            progress_placeholder.info(f"📊 Fase 2: Extrayendo página {page_num+1} ({idx+1}/{len(relevant_pages)})...")
        img_bytes = base64.standard_b64decode(images[page_num])
        page_text = extract_convenio_from_image(client, img_bytes, "png", detailed=True)
        all_text += f"\n--- PÁGINA {page_num+1} ---\n{page_text}\n"

    return all_text

def buscar_convenio_con_ia(client, nombre_convenio):
    """Busca información del convenio usando Claude con búsqueda web para obtener datos actualizados"""

    prompt = f"""Eres un experto en convenios colectivos españoles y legislación laboral.

TAREA CRÍTICA: Busca en internet el convenio colectivo MÁS RECIENTE y ACTUALIZADO:
"{nombre_convenio}"

INSTRUCCIONES DE BÚSQUEDA:
1. Busca PRIMERO en el BOE (Boletín Oficial del Estado) o boletines autonómicos/provinciales
2. Busca la ÚLTIMA revisión salarial o tablas salariales publicadas
3. Asegúrate de obtener los datos del año actual o el más reciente disponible
4. Si hay varias publicaciones, usa SIEMPRE la más reciente

DEBES INCLUIR CON DATOS ACTUALIZADOS:

1. **TABLAS SALARIALES VIGENTES** - Salarios base por categoría profesional (indicar año de las tablas)
2. **COMPLEMENTOS SALARIALES**:
   - Plus de transporte
   - Plus de nocturnidad
   - Plus de festividad
   - Otros pluses específicos del sector

3. **ANTIGÜEDAD**:
   - Tipo (trienios, quinquenios, bienios)
   - Porcentaje o cantidad por periodo

4. **PAGAS EXTRAORDINARIAS**:
   - Número de pagas extras
   - Cuantía (salario base, base + antigüedad, etc.)

5. **JORNADA LABORAL**:
   - Horas anuales
   - Horas semanales

6. **CATEGORÍAS PROFESIONALES** del sector

7. **FECHA DE PUBLICACIÓN** del convenio/tablas salariales encontradas

8. **FUENTE** (enlace al BOE u otra fuente oficial)

IMPORTANTE:
- NO uses datos aproximados ni de memoria
- BUSCA siempre en internet para obtener los datos más recientes
- Indica claramente la fecha y fuente de los datos

Responde de forma estructurada y detallada para poder calcular costes de subrogación.
"""

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=8192,
        tools=[
            {
                "type": "web_search_20250305"
            }
        ],
        messages=[
            {
                "role": "user",
                "content": prompt
            }
        ]
    )

    # Extraer el texto de la respuesta (puede venir en varios bloques por la búsqueda web)
    result_text = ""
    for block in response.content:
        if hasattr(block, 'text'):
            result_text += block.text + "\n"

    return result_text.strip()

def analyze_with_claude(client, file_bytes, file_type, convenio_text, years, is_image=False):
    """Analiza el documento con Claude"""

    prompt = f"""Eres un experto en recursos humanos y cálculo de costes de subrogación de personal en España.

TAREA: Analiza la tabla de personal del documento adjunto y calcula los costes de subrogación con PRECISIÓN.

CONVENIO DE REFERENCIA:
{convenio_text[:50000] if convenio_text else "No se ha proporcionado convenio de referencia."}

=== ANÁLISIS DEL CONVENIO - PRIMERO ===

Antes de calcular, EXTRAE del convenio estos datos (si no están, indica "No especificado"):

1. **NÚMERO DE PAGAS**: ¿Cuántas pagas al año? (12, 14, 15...)
2. **SALARIO BASE** por categoría profesional (mensual y anual)
3. **PLUSES Y COMPLEMENTOS** (indicar importe mensual de cada uno):
   - Plus Transporte / Locomoción
   - Plus Convenio / Plus Empresa
   - Plus Nocturnidad (% sobre salario base)
   - Plus Festivos / Domingos
   - Plus Toxicidad / Penosidad / Peligrosidad
   - Plus Disponibilidad
   - Plus Asistencia
   - Otros pluses específicos del sector
4. **ANTIGÜEDAD**:
   - Tipo: Trienios / Bienios / Quinquenios
   - Importe o % por cada periodo
   - ¿Se aplica sobre salario base o sobre total?
5. **PAGAS EXTRAORDINARIAS**:
   - Cuántas pagas extras
   - Base de cálculo (solo salario base, base+antigüedad, salario total...)

=== DATOS A EXTRAER DE LA TABLA DE PERSONAL ===

- Trabajador (nombre o iniciales)
- Antigüedad (fecha de alta)
- Tipo contrato (código o descripción)
- Categoría profesional
- **JORNADA MENSUAL** (en horas) - MUY IMPORTANTE
- Salario bruto anual (IGNORAR - recalcular desde convenio)

=== CÁLCULO DE JORNADA ===

**JORNADA COMPLETA DE REFERENCIA:**
- Jornada completa mensual = 152 horas/mes (aprox. 1.824 horas/año)
- Si un trabajador tiene 80 horas/mes = 52,6% de jornada
- Si un trabajador tiene 148 horas/mes = 97,4% de jornada

**PORCENTAJE DE JORNADA:**
% Jornada = (Horas mensuales del trabajador / 152) × 100

=== FÓRMULAS DE CÁLCULO ===

**IMPORTANTE: IGNORAR los salarios que aparezcan en la tabla (pueden estar desactualizados)**
**SIEMPRE calcular los salarios desde el CONVENIO proporcionalmente a la jornada**

**CÁLCULO DETALLADO DEL SALARIO MENSUAL:**
1. Salario Base (según categoría y convenio) × % Jornada
2. Plus Antigüedad = (Importe trienio × Nº Trienios) × % Jornada
3. Plus Transporte × % Jornada (si aplica)
4. Plus Convenio × % Jornada (si aplica)
5. Otros Pluses × % Jornada (si aplican)

**Salario Bruto Mensual** = Suma de todos los conceptos anteriores

**CÁLCULO ANUAL:**
- Si el convenio indica X pagas → Salario Bruto Anual = Salario Bruto Mensual × X pagas
- Las pagas extras pueden tener base de cálculo diferente (verificar en convenio)

**COSTE EMPRESA:**
- SS Empresa Anual = Salario Bruto Anual × 0.32
- COSTE EMPRESA ANUAL = Salario Bruto Anual + SS Empresa Anual

=== INSTRUCCIONES ===

1. **PRIMERO**: Muestra un resumen de los datos del convenio encontrados:

| Concepto | Valor según convenio |
|----------|---------------------|
| Número de pagas | X |
| Salario base [Categoría] | X €/mes |
| Plus Transporte | X €/mes |
| Plus Convenio | X €/mes |
| Antigüedad (trienio) | X €/mes o X% |
| Otros pluses | ... |

2. Extrae TODOS los trabajadores con sus datos

3. Para CADA trabajador calcula DESGLOSANDO todos los conceptos:
   - Años de antigüedad = Fecha actual - Fecha alta
   - Nº Trienios = Años antigüedad / 3 (parte entera)
   - % Jornada = Horas mensuales / 152 × 100
   - Salario Base × % Jornada = X €
   - Plus Antigüedad (X trienios × importe) × % Jornada = X €
   - Plus Transporte × % Jornada = X €
   - Plus Convenio × % Jornada = X €
   - Otros pluses × % Jornada = X €
   - **Total Mensual** = Suma
   - **Bruto Anual** = Total Mensual × Nº Pagas
   - SS Empresa = Bruto Anual × 0.32
   - **COSTE EMPRESA ANUAL** = Bruto Anual + SS Empresa

4. FACTORES ADICIONALES:
   - **Suplencia vacaciones**: 1 mes de suplencia = Coste mensual empresa
   - **Absentismo 2%**: (Coste personal + Suplencias) × 0.02

5. Período de cálculo: {years} año(s)

6. **TABLA DE PERSONAL DETALLADA:**
| Trabajador | Categoría | Trienios | % Jornada | Base | Antigüedad | Transporte | Otros Pluses | Total Mes | Bruto Anual | SS Empresa | Coste Empresa |

7. **TABLA RESUMEN DE COSTES:**

| Concepto | Año 1 | Total {years} Año(s) |
|----------|-------|----------------------|
| Coste Personal (suma todos) | € | € |
| Suplencia Vacaciones (1 mes) | € | € |
| Absentismo (2%) | € | € |
| **SUBTOTAL PERSONAL** | € | € |
| Gastos Generales (8%) | € | € |
| Materiales Estimados | € | € |
| **TOTAL GENERAL** | € | € |

8. **RESUMEN DE HORAS:**
| Concepto | Valor |
|----------|-------|
| Total trabajadores | X |
| Horas mensuales totales | X horas |
| Horas anuales totales | X horas |
| Equivalente jornadas completas | X |

9. **OBSERVACIONES**:
   - Indica qué pluses se han aplicado y cuáles no
   - Si algún dato no estaba en el convenio, indícalo
   - Cualquier observación relevante sobre el cálculo

IMPORTANTE:
- RESPETAR las horas de jornada de cada trabajador
- IGNORAR salarios de la tabla - calcular SIEMPRE desde el convenio
- Aplicar proporción de jornada a todos los conceptos salariales
- Calcular SS Empresa como ~32% del bruto
- Formato español: punto miles, coma decimales (18.456,78 €)
"""

    messages_content = []

    if is_image:
        img_base64 = base64.standard_b64encode(file_bytes).decode("utf-8")
        media_type = "image/png" if file_type == "png" else f"image/{file_type}"
        messages_content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": media_type,
                "data": img_base64,
            }
        })
    else:
        images = pdf_to_images(file_bytes)
        for img_base64 in images[:10]:
            messages_content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": "image/png",
                    "data": img_base64,
                }
            })

    messages_content.append({
        "type": "text",
        "text": prompt
    })

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[
            {
                "role": "user",
                "content": messages_content
            }
        ]
    )

    return response.content[0].text

def parse_markdown_tables(text):
    """Extrae tablas markdown del texto"""
    tables = []
    lines = text.split('\n')
    current_table = []
    in_table = False

    for line in lines:
        if '|' in line and line.strip().startswith('|'):
            in_table = True
            current_table.append(line)
        elif in_table and line.strip() == '':
            if current_table:
                tables.append(current_table)
                current_table = []
            in_table = False
        elif in_table and '|' not in line:
            if current_table:
                tables.append(current_table)
                current_table = []
            in_table = False

    if current_table:
        tables.append(current_table)

    return tables

def create_excel_from_result(resultado, years):
    """Crea un archivo Excel con los resultados del análisis"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Costes Subrogación"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    money_alignment = Alignment(horizontal='right')
    center_alignment = Alignment(horizontal='center')

    # Título
    ws['A1'] = f"CÁLCULO DE COSTES DE SUBROGACIÓN - {years} AÑO(S)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:I1')

    # Intentar extraer tablas del resultado
    tables = parse_markdown_tables(resultado)

    current_row = 3

    if tables:
        for table in tables:
            for line in table:
                # Limpiar la línea
                if '---' in line:
                    continue
                cells = [cell.strip() for cell in line.split('|') if cell.strip()]
                if cells:
                    for col, cell in enumerate(cells, 1):
                        ws.cell(row=current_row, column=col, value=cell)
                        ws.cell(row=current_row, column=col).border = border
                        if current_row == 3 or (table == tables[0] and line == table[0]):
                            ws.cell(row=current_row, column=col).font = header_font
                            ws.cell(row=current_row, column=col).fill = header_fill
                    current_row += 1
            current_row += 2
    else:
        # Si no hay tablas, poner el texto completo
        ws['A3'] = "Resultado del Análisis:"
        ws['A3'].font = Font(bold=True)
        current_row = 5
        for line in resultado.split('\n'):
            if line.strip():
                ws.cell(row=current_row, column=1, value=line)
                current_row += 1

    # Ajustar anchos de columna
    for col_num in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50) if max_length > 0 else 10
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar en buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer

def main():
    st.set_page_config(
        page_title="Calculadora de Subrogación",
        page_icon="📊",
        layout="wide"
    )

    # Verificar login
    if not check_login():
        return

    # Usuario autenticado - mostrar botón de logout en sidebar
    with st.sidebar:
        st.markdown(f"**Usuario:** {get_secret('LOGIN_USER')}")
        if st.button("🚪 Cerrar Sesión"):
            st.session_state.authenticated = False
            st.rerun()
        st.markdown("---")

    st.title("📊 Calculadora de Costes de Subrogación")
    st.markdown("---")

    # Verificar API key
    api_key = get_secret("ANTHROPIC_API_KEY")
    if not api_key:
        st.error("❌ No se encontró la API key de Anthropic.")
        st.stop()

    # Inicializar cliente de Anthropic
    client = anthropic.Anthropic(api_key=api_key)

    # Sidebar con configuración
    with st.sidebar:
        st.header("⚙️ Configuración")

        years = st.number_input(
            "Años para el cálculo",
            min_value=1,
            max_value=10,
            value=1,
            help="Número de años para calcular el coste total de subrogación"
        )

        st.markdown("---")

        st.header("📁 Convenio de Referencia")

        # Selector de método de convenio
        metodo_convenio = st.radio(
            "¿Cómo quieres indicar el convenio?",
            options=["Buscar con IA", "Seleccionar archivo", "Subir archivo"],
            help="La IA puede buscar información del convenio por su nombre"
        )

        convenio_seleccionado = "Ninguno"
        convenio_subido = None
        convenio_busqueda = ""

        if metodo_convenio == "Buscar con IA":
            convenio_busqueda = st.text_input(
                "Nombre del convenio",
                placeholder="Ej: Convenio colectivo de limpieza de Madrid",
                help="Escribe el nombre del convenio y la IA buscará la información"
            )
            if convenio_busqueda:
                st.success(f"✅ Se buscará en internet la versión más reciente: {convenio_busqueda}")

        elif metodo_convenio == "Seleccionar archivo":
            convenios = get_convenios_disponibles()
            if convenios:
                convenio_seleccionado = st.selectbox(
                    "Selecciona un convenio",
                    options=["Ninguno"] + [c.name for c in convenios]
                )
            else:
                st.info("No hay convenios PDF en la carpeta")

        else:  # Subir archivo
            convenio_subido = st.file_uploader(
                "Sube el convenio (PDF o imagen)",
                type=["pdf", "png", "jpg", "jpeg"],
                key="convenio_upload"
            )
            if convenio_subido and convenio_subido.type.startswith("image"):
                st.image(convenio_subido, caption="Preview del convenio", use_container_width=True)

    # Área principal
    col1, col2 = st.columns([1, 1])

    with col1:
        st.header("📄 Subir Tabla de Personal")

        uploaded_file = st.file_uploader(
            "Sube un PDF o imagen con la tabla de personal",
            type=["pdf", "png", "jpg", "jpeg"],
            help="Formatos soportados: PDF, PNG, JPG"
        )

        if uploaded_file:
            st.success(f"✅ Archivo cargado: {uploaded_file.name}")
            if uploaded_file.type.startswith("image"):
                st.image(uploaded_file, caption="Preview", use_container_width=True)

    with col2:
        st.header("📋 Información")

        # Determinar qué convenio mostrar
        if metodo_convenio == "Buscar con IA" and convenio_busqueda:
            convenio_mostrar = f"Buscar con IA: {convenio_busqueda}"
        elif convenio_seleccionado != "Ninguno":
            convenio_mostrar = convenio_seleccionado
        elif convenio_subido:
            convenio_mostrar = convenio_subido.name
        else:
            convenio_mostrar = "No seleccionado"

        st.info(f"""
        **Configuración actual:**
        - Años de cálculo: **{years}**
        - Convenio: **{convenio_mostrar}**
        """)

        st.markdown("""
        **La IA analizará:**
        - Datos de cada trabajador
        - Categoría profesional
        - Antigüedad
        - Salario y complementos
        - Costes de Seguridad Social
        """)

    st.markdown("---")

    # Botón de análisis
    if uploaded_file:
        if st.button("🔍 Analizar y Calcular Costes", type="primary", use_container_width=True):

            convenio_text = ""

            # Obtener información del convenio según el método seleccionado
            if metodo_convenio == "Buscar con IA" and convenio_busqueda:
                with st.spinner(f"🌐 Buscando en internet el convenio más reciente: {convenio_busqueda}..."):
                    convenio_text = buscar_convenio_con_ia(client, convenio_busqueda)
                    st.success("✅ Convenio actualizado obtenido de internet")
            elif convenio_subido:
                progress_placeholder = st.empty()
                progress_placeholder.info("📄 Procesando archivo del convenio...")
                convenio_bytes = convenio_subido.read()
                is_convenio_image = convenio_subido.type.startswith("image")
                convenio_file_type = convenio_subido.type.split("/")[-1]
                convenio_text = extract_convenio_from_file(client, convenio_bytes, convenio_file_type, is_convenio_image, progress_placeholder)
                convenio_subido.seek(0)
                progress_placeholder.success("✅ Convenio procesado correctamente")
            elif convenio_seleccionado != "Ninguno":
                carpeta = Path(__file__).parent
                convenio_path = carpeta / convenio_seleccionado
                with open(convenio_path, "rb") as f:
                    convenio_text = extract_text_from_pdf(f.read())

            with st.spinner("🔄 Analizando documento con IA... Esto puede tardar unos segundos."):
                try:
                    file_bytes = uploaded_file.read()
                    is_image = uploaded_file.type.startswith("image")
                    file_type = uploaded_file.type.split("/")[-1]

                    resultado = analyze_with_claude(
                        client,
                        file_bytes,
                        file_type,
                        convenio_text,
                        years,
                        is_image
                    )

                    # Guardar resultado en session_state
                    st.session_state.resultado = resultado
                    st.session_state.years = years

                    st.success("✅ Análisis completado")

                except Exception as e:
                    st.error(f"❌ Error al analizar: {str(e)}")
                    st.exception(e)

    # Mostrar resultados si existen
    if "resultado" in st.session_state:
        st.header("📊 Resultados del Análisis")
        st.markdown(st.session_state.resultado)

        st.markdown("---")
        st.subheader("📥 Descargar Resultados")

        col_download1, col_download2 = st.columns(2)

        with col_download1:
            # Descargar como texto
            st.download_button(
                label="📄 Descargar como TXT",
                data=st.session_state.resultado,
                file_name=f"subrogacion_{st.session_state.years}_anios.txt",
                mime="text/plain",
                use_container_width=True
            )

        with col_download2:
            # Descargar como Excel
            excel_buffer = create_excel_from_result(
                st.session_state.resultado,
                st.session_state.years
            )
            st.download_button(
                label="📊 Descargar como XLSX",
                data=excel_buffer,
                file_name=f"subrogacion_{st.session_state.years}_anios.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    else:
        st.info("👆 Sube un archivo PDF o imagen con la tabla de personal para comenzar")

    # Footer
    st.markdown("---")
    st.caption("Desarrollado para cálculo de costes de subrogación de personal | Powered by Claude AI")

if __name__ == "__main__":
    main()
